"""
XlsxHandler 单元测试
======================
在内存里构造真实 .xlsx blob，覆盖 extract / describe / render / 健壮性 / 注册。

# 人工编写
"""
import sys
import unittest
from io import BytesIO
from pathlib import Path

# 让 tests/ 也能跑（项目用相对 import 的方式跑 main，这里也对齐）
ROOT = Path(__file__).parent.parent / "code"
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook  # noqa: E402

from core.handlers import HandlerRegistry  # noqa: E402
from core.handlers.xlsx_handler import XlsxHandler  # noqa: E402  触发注册


def make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    """根据 {sheet_name: [[行..]]} 构造一个 .xlsx blob。"""
    wb = Workbook()
    # 删掉默认的第一个 sheet 后再按入参顺序加
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestExtractText(unittest.TestCase):
    def setUp(self):
        self.h = XlsxHandler()

    def test_extract_text_single_sheet(self):
        blob = make_xlsx({"Sheet1": [["name", "value"], ["foo", 42]]})
        out = self.h.extract_text(blob)
        self.assertIsNotNone(out)
        self.assertIn("=== Sheet: Sheet1 ===", out)
        self.assertIn("A1=name", out)
        self.assertIn("B1=value", out)
        self.assertIn("A2=foo", out)
        self.assertIn("B2=42", out)

    def test_extract_text_multi_sheets(self):
        blob = make_xlsx({
            "First": [["x"]],
            "Second": [["y", "z"]],
        })
        out = self.h.extract_text(blob)
        self.assertIn("=== Sheet: First ===", out)
        self.assertIn("=== Sheet: Second ===", out)
        # 顺序：First 应当出现在 Second 之前
        self.assertLess(out.index("First"), out.index("Second"))

    def test_extract_text_skips_empty_cells(self):
        # 行: [None, "kept", None]——只有 B 列有值
        blob = make_xlsx({"S": [[None, "kept", None]]})
        out = self.h.extract_text(blob)
        self.assertIn("B1=kept", out)
        self.assertNotIn("A1=", out)
        self.assertNotIn("C1=", out)


class TestDescribeChange(unittest.TestCase):
    def setUp(self):
        self.h = XlsxHandler()

    def test_describe_change_modify_cell(self):
        old = make_xlsx({"S": [["a", "b"], ["c", "d"]]})
        new = make_xlsx({"S": [["a", "b"], ["c", "D"]]})  # 仅 B2 改 d→D
        desc = self.h.describe_change(old, new)
        # 一个单元格变化 = +1/-1 = 2 行变更（这是 cell-level 的语义）
        # 期望表述为"修改 X 个单元格"
        self.assertIn("修改", desc)
        self.assertIn("单元格", desc)
        self.assertIn("1 个工作表", desc)

    def test_describe_change_add_cells(self):
        old = make_xlsx({"S": [["a"]]})
        new = make_xlsx({"S": [["a"], ["new1"], ["new2"]]})
        desc = self.h.describe_change(old, new)
        self.assertIn("修改", desc)
        self.assertIn("单元格", desc)

    def test_describe_change_no_content_diff(self):
        # 两个完全相同的工作簿——内容无差
        blob = make_xlsx({"S": [["a", "b"]]})
        # 重新序列化一份（字节可能不同但内容一样）
        blob2 = make_xlsx({"S": [["a", "b"]]})
        desc = self.h.describe_change(blob, blob2)
        # 两种合理结果：完全一致 / 仅元数据差异
        self.assertTrue(
            "仅元数据差异" in desc or "0 个单元格" in desc,
            f"unexpected desc: {desc!r}",
        )

    def test_describe_change_touches_multiple_sheets(self):
        old = make_xlsx({"A": [["x"]], "B": [["y"]]})
        new = make_xlsx({"A": [["X"]], "B": [["Y"]]})
        desc = self.h.describe_change(old, new)
        self.assertIn("2 个工作表", desc)


class TestRenderDiff(unittest.TestCase):
    def setUp(self):
        self.h = XlsxHandler()

    def test_render_diff_tags(self):
        old = make_xlsx({"S": [["a", "b"], ["c", "d"]]})
        new = make_xlsx({"S": [["a", "b"], ["c", "D"]]})
        out = self.h.render_diff(old, new)
        tags = {t for t, _ in out}
        self.assertIn("added", tags)
        self.assertIn("removed", tags)
        # sheet 头应当标为 meta
        self.assertIn("meta", tags)
        # 不应出现 difflib 的 '? ' hint
        self.assertNotIn("?", tags)
        # 内容应当包含 d 与 D
        contents = " ".join(c for _, c in out)
        self.assertIn("d", contents)
        self.assertIn("D", contents)

    def test_render_diff_sheet_header_is_meta(self):
        blob = make_xlsx({"S": [["a"]]})
        out = self.h.render_diff(blob, blob)
        # 同样的内容：sheet 头是 meta，其余是 normal
        sheet_lines = [(t, c) for t, c in out if c.startswith("=== Sheet:")]
        self.assertTrue(len(sheet_lines) >= 1)
        for t, _ in sheet_lines:
            self.assertEqual(t, "meta")

    def test_render_diff_has_content_lines(self):
        old = make_xlsx({"S": [["only_old"]]})
        new = make_xlsx({"S": [["only_new"]]})
        out = self.h.render_diff(old, new)
        # 至少要有一个 added 一个 removed
        added = [c for t, c in out if t == "added"]
        removed = [c for t, c in out if t == "removed"]
        self.assertTrue(any("only_new" in c for c in added))
        self.assertTrue(any("only_old" in c for c in removed))


class TestRegistry(unittest.TestCase):
    def test_xlsx_extension_routes_to_xlsx_handler(self):
        h = HandlerRegistry.for_path(Path("report.xlsx"))
        self.assertIsInstance(h, XlsxHandler)

    def test_xlsx_extension_case_insensitive(self):
        h = HandlerRegistry.for_path(Path("Q4.XLSX"))
        self.assertIsInstance(h, XlsxHandler)

    def test_known_extensions_includes_xlsx(self):
        self.assertIn(".xlsx", HandlerRegistry.known_extensions())


class TestRobustness(unittest.TestCase):
    def setUp(self):
        self.h = XlsxHandler()

    def test_extract_text_corrupt_blob_returns_none(self):
        # "not a zip"——openpyxl 会抛 BadZipFile，handler 必须吞掉
        out = self.h.extract_text(b"this is not a real xlsx file")
        self.assertIsNone(out)

    def test_describe_change_corrupt_blob_does_not_crash(self):
        good = make_xlsx({"S": [["a"]]})
        desc = self.h.describe_change(b"garbage", good)
        # 不抛异常即可；应当返回带"解析失败"提示的字符串
        self.assertIsInstance(desc, str)
        self.assertIn("解析失败", desc)

    def test_render_diff_corrupt_blob_returns_meta(self):
        out = self.h.render_diff(b"junk", b"more junk")
        # 至少返回一条 meta 提示，不能抛
        self.assertTrue(len(out) >= 1)
        self.assertEqual(out[0][0], "meta")


class TestFormulaTracking(unittest.TestCase):
    """追踪盲区修复：改公式必须可见（原版 data_only=True 会漏）。"""

    def setUp(self):
        self.h = XlsxHandler()

    @staticmethod
    def _xlsx_with_formula(formula: str) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = formula
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_formula_cell_extracted(self):
        # 公式单元格不能因为没有缓存值就被当成空格跳过
        out = self.h.extract_text(self._xlsx_with_formula("=A1+A2"))
        self.assertIn("=A1+A2", out)

    def test_formula_change_is_visible(self):
        old = self._xlsx_with_formula("=A1+A2")
        new = self._xlsx_with_formula("=A1*A2")
        desc = self.h.describe_change(old, new)
        self.assertNotIn("一致", desc)
        self.assertNotIn("仅元数据", desc)
        contents = " ".join(c for _, c in self.h.render_diff(old, new))
        self.assertIn("=A1+A2", contents)
        self.assertIn("=A1*A2", contents)


if __name__ == "__main__":
    unittest.main(verbosity=2)
