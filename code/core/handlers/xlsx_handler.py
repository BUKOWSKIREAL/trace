"""
XlsxHandler — .xlsx 表格文件的 diff / 摘要
==============================================
处理 Excel 工作簿：把每个工作表的非空单元格拍平成
"=== Sheet: {名} ===" 段 + "{坐标}={值}" 行的可读文本，
然后用行级 diff 给出单元格级别的变更摘要与可视化。

# AI生成 (Claude 4.7 协助)
# 关键选择：
#   - read_only=True + data_only=True：流式读取省内存、只取计算结果不取公式
#   - 跳过空单元格：减少噪声，让 diff 聚焦真实变更
#   - 按 row→col 排序：openpyxl 默认迭代顺序已是行优先；这里显式确认
#   - 出现 BadZipFile / InvalidFileException 等损坏情况要 catch 住，
#     不让异常冒泡到 UI——返回一条 meta 行即可
"""
from __future__ import annotations

import difflib
import logging
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.handlers.base import DiffLine, FileHandler, HandlerRegistry

logger = logging.getLogger("trace.handlers.xlsx")


class XlsxHandler(FileHandler):
    """Excel .xlsx handler，按"单元格"为语义单位做 diff。"""

    extensions = [".xlsx"]

    def extract_text(self, blob: bytes) -> str | None:
        """
        把工作簿拍平成确定性文本：每个 sheet 一段、每个非空单元格一行。
        损坏时返回 None——让上层走兜底显示，不让 UI 崩。
        """
        try:
            return self._flatten(blob)
        except Exception as exc:  # noqa: BLE001  外部库错误种类多，统一兜
            logger.warning("xlsx 解析失败: %s", exc)
            return None

    def describe_change(self, old_blob: bytes, new_blob: bytes) -> str:
        """
        摘要："修改 X 个单元格（涉及 Y 个工作表）" / 或解析失败提示。
        # 用行级 diff 数变更行数——每行就是一个单元格条目，所以行数=单元格数。
        # 顺便从 "=== Sheet: 名 ===" 行抽出涉及的 sheet 名做去重计数。
        """
        old_text = self._try_flatten(old_blob)
        new_text = self._try_flatten(new_blob)
        if old_text is None or new_text is None:
            return "[文件解析失败：无法对比]"

        old_lines = old_text.splitlines()
        new_lines = new_text.splitlines()

        changed_cells = 0
        sheets_touched: set[str] = set()
        current_sheet = ""

        for line in difflib.ndiff(old_lines, new_lines):
            tag2 = line[:2]
            content = line[2:]
            # sheet 头出现在 normal/added/removed 任一类型里都要更新当前 sheet 上下文
            if content.startswith("=== Sheet: ") and content.endswith(" ==="):
                current_sheet = content[len("=== Sheet: "):-len(" ===")]
                continue
            if tag2 in ("+ ", "- "):
                changed_cells += 1
                if current_sheet:
                    sheets_touched.add(current_sheet)

        if changed_cells == 0:
            return "仅元数据差异（单元格内容一致）"
        return f"修改 {changed_cells} 个单元格（涉及 {len(sheets_touched)} 个工作表）"

    def render_diff(self, old_blob: bytes, new_blob: bytes) -> list[DiffLine]:
        """
        生成 (tag, content) 列表。'=== Sheet: ... ===' 始终标 meta；
        其余按 ndiff 标 normal/added/removed；'? ' hint 行跳过。
        解析失败时返回单行 meta，UI 友好提示。
        """
        if not old_blob and not new_blob:
            return []

        if not old_blob:
            new_text = self._try_flatten(new_blob)
            if new_text is None:
                return [("meta", "[文件解析失败：新版本无法解析为 xlsx]")]
            return self._text_as_diff(new_text, "added")

        if not new_blob:
            old_text = self._try_flatten(old_blob)
            if old_text is None:
                return [("meta", "[文件解析失败：旧版本无法解析为 xlsx]")]
            return self._text_as_diff(old_text, "removed")

        old_text = self._try_flatten(old_blob)
        new_text = self._try_flatten(new_blob)
        if old_text is None and new_text is None:
            return [("meta", "[文件解析失败：新旧版本均无法解析为 xlsx]")]
        if old_text is None:
            return [("meta", "[文件解析失败：旧版本无法解析为 xlsx]")]
        if new_text is None:
            return [("meta", "[文件解析失败：新版本无法解析为 xlsx]")]

        old_lines = old_text.splitlines()
        new_lines = new_text.splitlines()

        result: list[DiffLine] = []
        for line in difflib.ndiff(old_lines, new_lines):
            tag2 = line[:2]
            content = line[2:]
            # sheet 标题行无论新增/删除/不变都用 meta 标——它不是真正的"数据变更"
            is_sheet_header = content.startswith("=== Sheet: ") and content.endswith(" ===")
            if is_sheet_header:
                result.append(("meta", content))
                continue
            if tag2 == "  ":
                result.append(("normal", content))
            elif tag2 == "+ ":
                result.append(("added", content))
            elif tag2 == "- ":
                result.append(("removed", content))
            # '? ' 跳过
        return result

    # ---------- 内部工具 ----------

    def _try_flatten(self, blob: bytes) -> str | None:
        """_flatten 的安全版本——失败返回 None 而非抛异常。"""
        try:
            return self._flatten(blob)
        except Exception as exc:  # noqa: BLE001
            logger.warning("xlsx 解析失败: %s", exc)
            return None

    def _flatten(self, blob: bytes) -> str:
        """
        真正的拍平逻辑。
        # read_only 模式下 sheet 对象没有 .dimensions，但 iter_rows 仍可用；
        # values_only=True 只拿值不构 Cell 实例，省内存。
        #
        # 人工修正（追踪盲区修复）：原版只用 data_only=True 读"缓存的计算结果"，
        # 导致两个问题——
        #   1. 改公式（=A1+A2 → =A1*A2）但缓存值未刷新时，diff 看不到任何变化；
        #   2. openpyxl 在没有缓存值时对公式单元格返回 None，整行被当成空单元格跳过，
        #      用户明明写了公式却显示"该格为空"。
        # 修法：读两遍——data_only=False 拿公式文本，data_only=True 拿缓存值，
        # 公式单元格输出 "坐标=公式 (值)"，普通单元格仍输出 "坐标=值"。
        """
        # 第一遍：拿公式 / 原始内容（data_only=False）
        wb_f = load_workbook(BytesIO(blob), data_only=False, read_only=True)
        # 第二遍：拿缓存的计算结果（data_only=True）
        wb_v = load_workbook(BytesIO(blob), data_only=True, read_only=True)
        try:
            lines: list[str] = []
            for sheet_name in wb_f.sheetnames:
                lines.append(f"=== Sheet: {sheet_name} ===")
                ws_f = wb_f[sheet_name]
                ws_v = wb_v[sheet_name] if sheet_name in wb_v.sheetnames else None
                rows_v = (
                    ws_v.iter_rows(values_only=True) if ws_v is not None else iter([])
                )
                # 两个工作簿同结构，逐行 zip；值侧缺行用空元组兜
                for r_idx, row_f in enumerate(
                    ws_f.iter_rows(values_only=True), start=1
                ):
                    row_v = next(rows_v, ())
                    for c_idx, raw in enumerate(row_f, start=1):
                        if raw is None:
                            continue
                        coord = f"{get_column_letter(c_idx)}{r_idx}"
                        # 公式单元格：openpyxl 在 data_only=False 下返回以 '=' 开头的字符串
                        is_formula = isinstance(raw, str) and raw.startswith("=")
                        if is_formula:
                            cached = row_v[c_idx - 1] if c_idx - 1 < len(row_v) else None
                            if cached is not None:
                                lines.append(f"{coord}={raw} (值: {cached})")
                            else:
                                lines.append(f"{coord}={raw}")
                        else:
                            lines.append(f"{coord}={raw}")
            return "\n".join(lines)
        finally:
            # read_only 模式持有文件句柄/zip 引用，显式关闭释放
            wb_f.close()
            wb_v.close()


# 模块导入时自动注册
HandlerRegistry.register(XlsxHandler())
